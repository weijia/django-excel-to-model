from openpyxl import Workbook, load_workbook

from django_excel_to_model.file_readers.data_source_base import DataSourceBase, RowRandomlyAccessibleSheet
from django_excel_to_model.reader import BaseSheet


# Ref: https://www.osgeo.cn/openpyxl/tutorial.html
# Ref: https://zhuanlan.zhihu.com/p/51292549
class XlsxSheetWithOpenOpenpyxl(RowRandomlyAccessibleSheet):
    def __init__(self, workbook, index):
        super(XlsxSheetWithOpenOpenpyxl, self).__init__()
        self.headers = None
        self.sheet = workbook[workbook.sheetnames[index]]
        self.header_row_start_from_0 = None

    def enumerate(self, start_row_numbered_from_0=1):
        for row in self.sheet.rows[start_row_numbered_from_0:]:
            row_data = self.raw_data_to_row(row)
            if row_data.count(None) == len(row_data):
                continue
            yield row_data

    def get_row(self, row_index_start_from_0):
        return self.raw_data_to_row(self.sheet.rows[row_index_start_from_0])

    def raw_data_to_row(self, raw_row_data):
        """

        :param raw_row_data: raw data read from openpyxl.
                        column.has_style    column.number_format    value type
              date:     True                'mm-dd-yy'              datetime
              ?float:   True                '0.00'                  float
              int:      True                '0.00'                  long
              str:      False               '0.00'                  unicode
        :return:
        """
        row = []
        for column in raw_row_data:
            row.append(column.value)
        return row

    def get_total_rows(self):
        return len(self.sheet.rows)


class OpenpyxlExcelFile(object):
    def __init__(self, full_path, flag=None):
        super(OpenpyxlExcelFile, self).__init__()
        self.wb = load_workbook(full_path)

    def get_sheet(self, sheet_index_numbered_from_0):
        return XlsxSheetWithOpenOpenpyxl(self.wb, sheet_index_numbered_from_0)
